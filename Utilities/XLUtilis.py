import openpyxl

# filename = r"C:\Users\Admin\PycharmProjects\seleniumDemo\Excel.xlsx"
# # Load/ open excel file by name
# workbook = openpyxl.load_workbook(filename)
#
# # Locate /get perticular sheet
# sheet = workbook['Sheet1']              ## UserID is sheetname, e.g. sheet1,sheet2
#
# # Read data from sheet
# #
# # print(sheet.cell(row=1, column=1).value)
# # print(sheet.cell(row=1, column=2).value)
#
# # Assignment : Read all values from excel
# sheet.cell(row=9, column=3).value = "Shreya"
# # sheet.cell(row=9, column=1).value = "Batch 12"
# # sheet.cell(row=9, column=2).value = "Batch_12_Pass"
# # #
# # print(sheet.cell(row=9, column=1).value + " " + sheet.cell(row=9, column=2).value)
# # #
# workbook.save(filename)

def getRowCount(file,sheetName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return(sheet.max_row)

def getColumnCount(file,sheetName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return(sheet.max_column)

def readData(file,sheetName,rownum,columno):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return sheet.cell(row=rownum,column=columno).value
#
# def writeData(file,sheetName,rownum,columno,data):
#     workbook = openpyxl.load_workbook(file)
#     sheet = workbook[sheetName]
#     return sheet.cell(row=rownum,column=columno).value = data
#     workbook.save(file)


